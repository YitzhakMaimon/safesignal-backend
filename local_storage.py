import os
import threading
from datetime import datetime

from openpyxl import Workbook, load_workbook

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

ERROR_LOG_PATH = os.path.join(DATA_DIR, "error_log.xlsx")
ERROR_LOG_SHEET_NAME = "Errors"
ERROR_LOG_COLUMNS = [
    "timestamp", "source", "incident_id", "user_id", "error_message",
]

FALSE_POSITIVE_LOG_PATH = os.path.join(DATA_DIR, "false_positive_log.xlsx")
FALSE_POSITIVE_LOG_SHEET_NAME = "FalsePositives"
FALSE_POSITIVE_LOG_COLUMNS = [
    "timestamp", "incident_id", "user_id", "platform", "raw_text",
    "risk_level", "distress_classification", "summary",
]

TOKENS_LOG_PATH = os.path.join(DATA_DIR, "tokens_log.xlsx")
TOKENS_LOG_SHEET_NAME = "TokenUsage"

# One column per model/pipeline-stage that consumes real tokens anywhere in the
# pipeline, from n8n's very first model call (Image Analysis / Gemini Vision,
# at ingestion) through the last one (the Decision Agent's post-hallucination-
# check pass) -- ml_comprehend.py, distress_classification.py, info_extraction.py,
# decision_agent_graph.py, safesignal.py's _analyze_image. An incident gets ONE
# row here no matter how many of these it passes through in a full run;
# whichever ones it never reached stay 0.
#
# Groq Whisper (voice transcription, also at ingestion) is deliberately NOT a
# column here: it's billed by audio duration, not tokens, and the Groq SDK's
# Transcription response has no usage/token field to report (verified against
# groq.types.audio.transcription.Transcription) -- there is nothing real to log,
# so this stays an honest gap rather than a fabricated number.
TOKENS_LOG_MODEL_COLUMNS = [
    "HeBERT_screening_tokens",
    "Bedrock_Classification_tokens",
    "Bedrock_Translation_CrossCheck_tokens",
    "Ollama_Extraction_tokens",
    "Gemini_DecisionAgent_tokens",
    "Bedrock_HallucinationCheck_tokens",
    "Gemini_Vision_tokens",
]
TOKENS_LOG_COLUMNS = ["timestamp", "incident_id", "sentence", *TOKENS_LOG_MODEL_COLUMNS, "total_tokens"]

# Maps each call site's pipeline_stage string to its column above. Both Gemini calls
# inside decision_agent_node (tool-check pass + structured-output pass) share one
# column -- same model, same node -- and add together rather than overwrite.
_STAGE_TO_TOKENS_COLUMN = {
    "ml_comprehend_hebert_screening": "HeBERT_screening_tokens",
    "distress_classification": "Bedrock_Classification_tokens",
    "ml_comprehend_translation_crosscheck": "Bedrock_Translation_CrossCheck_tokens",
    "info_extraction": "Ollama_Extraction_tokens",
    "decision_agent_tool_check": "Gemini_DecisionAgent_tokens",
    "decision_agent_structured_output": "Gemini_DecisionAgent_tokens",
    "decision_agent_hallucination_check": "Bedrock_HallucinationCheck_tokens",
    "vision_gemini_analysis": "Gemini_Vision_tokens",
}

_lock = threading.Lock()


def _ensure_tokens_log_schema(ws) -> bool:
    """
    Adds any TOKENS_LOG_COLUMNS entry missing from row 1 (e.g. a new model
    column introduced after tokens_log.xlsx already had rows on disk),
    inserting each one directly before its current position -- openpyxl's
    insert_cols shifts every existing cell after that point right by one,
    so already-written rows (including total_tokens) keep their real values
    lined up under the right header instead of reading off-by-one once the
    column list grows. Returns True if the sheet was changed (caller saves).
    """
    header = [cell.value for cell in ws[1]]
    changed = False
    for i, name in enumerate(TOKENS_LOG_COLUMNS):
        if i < len(header) and header[i] == name:
            continue
        ws.insert_cols(i + 1)
        ws.cell(row=1, column=i + 1, value=name)
        header.insert(i, name)
        changed = True
    return changed


def _append_row(path: str, sheet_name: str, columns: list[str], row: list) -> None:
    """
    Shared append-with-init helper for all local Excel-backed tables in this
    module. Kept local-only (no S3/RDS) by design -- this data is sensitive
    enough that it stays on disk, not in the cloud, until a real encrypted
    store replaces it.
    """
    with _lock:
        if not os.path.exists(path):
            os.makedirs(os.path.dirname(path), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(columns)
            wb.save(path)

        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws.append(row)
        wb.save(path)


def save_error_record(
    source: str,
    error_message: str,
    incident_id: str = "",
    user_id: str = "",
    path: str = ERROR_LOG_PATH,
) -> None:
    """
    רשומת שגיאה מקומית (data/error_log.xlsx) - אותו דפוס בדיוק כמו
    review_queue/immediate_alert. יעד: מסך שגיאות עתידי ב-UI שיקרא מכאן
    דרך GET /api/v1/errors, כדי שכשל שרת (למשל n8n שלא הצליח להגיע ל-
    backend) לא ייבלע בשקט אלא יופיע למשתמש.
    """
    _append_row(path, ERROR_LOG_SHEET_NAME, ERROR_LOG_COLUMNS, [
        datetime.now().isoformat(timespec="seconds"),
        source,
        incident_id,
        user_id,
        error_message,
    ])


def get_error_records(path: str = ERROR_LOG_PATH) -> list[dict]:
    """קורא את data/error_log.xlsx בחזרה כרשימת dict-ים עבור מסך השגיאות ב-UI."""
    if not os.path.exists(path):
        return []

    with _lock:
        wb = load_workbook(path)
        ws = wb[ERROR_LOG_SHEET_NAME] if ERROR_LOG_SHEET_NAME in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

    return [dict(zip(ERROR_LOG_COLUMNS, row)) for row in rows]


def save_false_positive_record(
    incident_id: str,
    user_id: str = "",
    platform: str = "",
    raw_text: str = "",
    risk_level: str = "",
    distress_classification: str = "",
    summary: str = "",
    path: str = FALSE_POSITIVE_LOG_PATH,
) -> None:
    """
    רשומת False Positive מקומית (data/false_positive_log.xlsx) - אותו דפוס
    בדיוק כמו error_log. נכתב מ-update_incident_status ב-safesignal.py כשמפעיל
    המערכת מסמן אירוע כ-False Positive, כדי לשמור תיעוד מקומי (מזהה אירוע,
    הטקסט המקורי ועוד פרטי הסיווג) לצורך בדיקה/שיפור המודל בהמשך.
    """
    _append_row(path, FALSE_POSITIVE_LOG_SHEET_NAME, FALSE_POSITIVE_LOG_COLUMNS, [
        datetime.now().isoformat(timespec="seconds"),
        incident_id,
        user_id,
        platform,
        raw_text,
        risk_level,
        distress_classification,
        summary,
    ])


def save_token_usage_record(
    pipeline_stage: str,
    model_id: str,
    sentence: str,
    input_tokens: int = 0,
    output_tokens: int = 0,
    incident_id: str = "",
    path: str = TOKENS_LOG_PATH,
) -> None:
    """
    רשומת צריכת טוקנים מקומית (data/tokens_log.xlsx), נכתבת מכל אחת מנקודות
    הקריאה ל-LLM/מודל בפייפליין (ml_comprehend.py, distress_classification.py,
    info_extraction.py, decision_agent_graph.py, safesignal.py's _analyze_image),
    בין אם הריצה הגיעה מתנועה אמיתית דרך n8n/safesignal.py ובין אם מהרצה ידנית
    של בלוק הבדיקה (__main__) של אחד המודולים.

    בניגוד ל-error_log/false_positive_log (append-only, שורה חדשה בכל קריאה),
    כאן יש שורה אחת בלבד לכל אירוע. ההתאמה היא לפי incident_id כשהוא קיים
    (כי incident_id עובר לאורך כל שרשרת ה-n8n, מהכניסה -- Image Analysis/Voice
    Transcription -- ועד היציאה, גם כשה"משפט" עצמו משתנה בין שלבים, למשל טקסט
    OCR/תמלול מול הטקסט הסופי הממוזג), ורק כשאין incident_id (הרצה ידנית של
    __main__ בלי אחד) עוברים להתאמה לפי טקסט מדויק כמו קודם. כך משפט/אירוע
    שעובר כמה שלבים בפייפליין מסתכם בשורה אחת עם עמודה לכל מודל, ולא מתפצל
    לכמה שורות.
    """
    column = _STAGE_TO_TOKENS_COLUMN.get(pipeline_stage)
    if column is None:
        raise ValueError(f"Unknown pipeline_stage for token logging: {pipeline_stage!r}")
    tokens = input_tokens + output_tokens

    with _lock:
        if not os.path.exists(path):
            os.makedirs(os.path.dirname(path), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = TOKENS_LOG_SHEET_NAME
            ws.append(TOKENS_LOG_COLUMNS)
            wb.save(path)

        wb = load_workbook(path)
        ws = wb[TOKENS_LOG_SHEET_NAME] if TOKENS_LOG_SHEET_NAME in wb.sheetnames else wb.active
        if _ensure_tokens_log_schema(ws):
            wb.save(path)

        sentence_idx = TOKENS_LOG_COLUMNS.index("sentence")
        column_idx = TOKENS_LOG_COLUMNS.index(column)
        incident_idx = TOKENS_LOG_COLUMNS.index("incident_id")
        timestamp_idx = TOKENS_LOG_COLUMNS.index("timestamp")
        total_idx = TOKENS_LOG_COLUMNS.index("total_tokens")

        if incident_id:
            target_row = next(
                (row for row in ws.iter_rows(min_row=2) if row[incident_idx].value == incident_id),
                None,
            )
        else:
            target_row = next(
                (row for row in ws.iter_rows(min_row=2) if row[sentence_idx].value == sentence),
                None,
            )
        if target_row is None:
            blank_row = ["" if c in ("timestamp", "incident_id", "sentence") else 0 for c in TOKENS_LOG_COLUMNS]
            ws.append(blank_row)
            target_row = ws[ws.max_row]

        target_row[timestamp_idx].value = datetime.now().isoformat(timespec="seconds")
        target_row[incident_idx].value = incident_id or target_row[incident_idx].value
        target_row[sentence_idx].value = sentence
        target_row[column_idx].value = (target_row[column_idx].value or 0) + tokens
        target_row[total_idx].value = sum(
            (target_row[TOKENS_LOG_COLUMNS.index(c)].value or 0) for c in TOKENS_LOG_MODEL_COLUMNS
        )

        wb.save(path)


def get_total_tokens_for_sentence(
    sentence: str, incident_id: str = "", path: str = TOKENS_LOG_PATH
) -> int | None:
    """
    Reads back the total_tokens column of tokens_log.xlsx for the given
    incident (same incident_id-first, sentence-fallback match
    save_token_usage_record uses to find its row) -- called once the full
    pipeline has finished for an incident, so every stage that ran for it has
    already added its tokens in, from Image Analysis/Voice Transcription at
    the very start of the n8n chain through the Decision Agent's own
    hallucination check at the end.

    Returns None (not 0) when the log doesn't exist yet or the incident has
    no row -- the dashboard's History screen renders that as a blank cell
    rather than a misleading "0 tokens used", which no real incident has.

    Opens the workbook normally (not read_only) rather than the older
    read_only mode, because it needs to run the same schema migration
    save_token_usage_record does (_ensure_tokens_log_schema) -- otherwise a
    file written before a new model column existed would misalign every
    index computed from today's TOKENS_LOG_COLUMNS.
    """
    if not os.path.exists(path):
        return None

    with _lock:
        wb = load_workbook(path)
        ws = wb[TOKENS_LOG_SHEET_NAME] if TOKENS_LOG_SHEET_NAME in wb.sheetnames else wb.active
        if _ensure_tokens_log_schema(ws):
            wb.save(path)

        sentence_idx = TOKENS_LOG_COLUMNS.index("sentence")
        incident_idx = TOKENS_LOG_COLUMNS.index("incident_id")
        total_idx = TOKENS_LOG_COLUMNS.index("total_tokens")

        fallback = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if incident_id and row[incident_idx] == incident_id:
                return int(row[total_idx] or 0)
            if fallback is None and row[sentence_idx] == sentence:
                fallback = int(row[total_idx] or 0)

    return fallback
