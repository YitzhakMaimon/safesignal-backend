"""
LangGraph Decision Agent node + graph -- extracted from decision_agent_graph.py
into a standalone microservice (Layer 3, Service 4).

Loads tools from the two running FastMCP servers (alert_mcp.py on :8011,
rag_mcp.py on :8012 -- moved off :8001/:8002 in this same refactor to avoid
colliding with services/rag_service and services/image_analyser), binds them
to Gemini, and runs the full bottom half of the pipeline diagram:

    decision_agent -> [tool call needed?] -> execute_tools -> decision_agent
    decision_agent -> output_screening
        -- blocked/hallucinated, retries left  --> back to decision_agent
        -- blocked/hallucinated, retries used up --> human_review
        -- passed                                --> routing_by_risk_level
    routing_by_risk_level
        -- risk_level == "high"   --> immediate_alert -> END
        -- risk_level == "medium" --> human_review -> END
        -- risk_level == "low"    --> log_and_close -> END

ARCHITECTURAL CHANGE from the monolith version: output_screening_node no
longer runs AWS Bedrock Guardrails / Claude Haiku hallucination checks
in-process. Both checks now live in the standalone services/output_screening
microservice (async, T3.micro-optimized) and are reached over HTTP via
_check_output_via_screening_service(). This file no longer imports
`anthropic` or talks to Bedrock Guardrails/Claude Haiku directly at all --
that's services/output_screening's job now.

Prerequisites: alert_mcp.py and rag_mcp.py must already be running (unless
using use_mock_tools=True), GOOGLE_API_KEY must be set, and
services/output_screening must be reachable (a hard 3s client-side timeout
means an unreachable screening service degrades the run to human_review
rather than hanging or crashing -- see _check_output_via_screening_service).
"""
import io
import os
from contextlib import AsyncExitStack
from datetime import datetime

import boto3
import httpx
import requests
from botocore.exceptions import BotoCoreError, ClientError
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()

from langchain_core.messages import HumanMessage, SystemMessage, ToolMessage
from langchain_core.tools import tool as lc_tool
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_mcp_adapters.tools import load_mcp_tools
from langgraph.graph import END, START, StateGraph
from langgraph.prebuilt import ToolNode
from mcp import ClientSession
from mcp.client.streamable_http import streamablehttp_client

from local_storage import save_immediate_alert_record
from schemas import AgentState, DecisionOutput

ALERT_MCP_URL = os.environ.get("ALERT_MCP_URL", "http://localhost:8011/mcp")
RAG_MCP_URL = os.environ.get("RAG_MCP_URL", "http://localhost:8012/mcp")

MODEL_NAME = os.environ.get("DECISION_AGENT_MODEL", "gemini-2.5-flash")

# --- Output screening: now an HTTP call to services/output_screening ---
OUTPUT_SCREENING_URL = os.environ.get("OUTPUT_SCREENING_URL", "http://localhost:8005/v1/screen")
MAX_OUTPUT_RETRIES = 3

# --- Routing by risk level (n8n workflow, with a local fallback) ---
ROUTING_WEBHOOK_URL = os.environ.get("ROUTING_WEBHOOK_URL", "http://localhost:5678/webhook/risk-routing")
URGENCY_TO_RISK_LEVEL = {"critical": "high", "high": "high", "medium": "medium", "low": "low"}

BEDROCK_AWS_REGION = os.environ.get("BEDROCK_AWS_REGION", "us-east-1")

# --- Terminal nodes ---
ALERT_WEBHOOK_URL = os.environ.get("ALERT_WEBHOOK_URL", "http://localhost:8000/api/v1/immediate-alert")
REVIEW_QUEUE_URL = os.environ.get("REVIEW_QUEUE_URL", "http://localhost:8000/api/v1/review-queue")

S3_LOG_BUCKET = os.environ.get("SAFESIGNAL_S3_BUCKET", "")


async def load_all_mcp_tools(exit_stack: AsyncExitStack) -> list:
    """Connects to both FastMCP servers and returns their combined tool list."""
    tools = []
    for url in (ALERT_MCP_URL, RAG_MCP_URL):
        read, write, _ = await exit_stack.enter_async_context(streamablehttp_client(url))
        session = await exit_stack.enter_async_context(ClientSession(read, write))
        await session.initialize()
        tools.extend(await load_mcp_tools(session))
    return tools


def build_mock_tools() -> list:
    """Local stand-ins for the two MCP tools, with no network I/O."""

    @lc_tool
    def trigger_immediate_alert(incident_id: str, urgency_reason: str) -> str:
        """
        Trigger an immediate human/Amazon Polly voice alert for a high-risk incident.
        Call this only when the situation requires urgent human intervention (e.g.
        Critical or High urgency with an imminent safety risk).

        Args:
            incident_id: Unique identifier of the incident being escalated.
            urgency_reason: Short factual explanation of why immediate escalation is required.
        """
        return f"[MOCK] ALERT_SENT: simulated immediate alert for incident '{incident_id}' ({urgency_reason})."

    @lc_tool
    def query_rag_history(user_id: str, target_query: str) -> dict:
        """
        Fetch a specific user's historical distress patterns and past incident logs.
        Call this when the user's own history -- not just similar cases from the
        general knowledge base -- would materially change the urgency assessment.

        Args:
            user_id: Identifier of the user whose history should be searched.
            target_query: The specific question or topic to search the user's history for.
        """
        return {
            "mock": True,
            "user_id": user_id,
            "target_query": target_query,
            "history": ["[MOCK] No real history store connected -- this is simulated data."],
        }

    return [trigger_immediate_alert, query_rag_history]


def _build_system_prompt(state: AgentState) -> str:
    return (
        "You are the Decision Agent in SafeSignal, an emergency and distress triage "
        "system. Evaluate the incident context below and decide whether an external "
        "tool must be called before you can produce a final assessment.\n\n"
        f"Incident ID: {state.incident_id}\n"
        f"User ID: {state.user_id}\n"
        f"Raw input: {state.raw_input}\n"
        f"Automated distress classification: {state.distress_classification}\n"
        f"Passive RAG context already retrieved: {state.initial_rag_context}\n\n"
        "Available tools:\n"
        "- trigger_immediate_alert: use ONLY for Critical/High urgency situations that "
        "need immediate human/Amazon Polly voice intervention.\n"
        "- query_rag_history: use when this specific user's own historical incident "
        "pattern would materially change the urgency assessment and isn't already "
        "covered by the passive RAG context above.\n\n"
        "If no tool is needed, respond directly -- your final answer will be parsed "
        "into a structured assessment."
    )


def make_decision_agent_node(llm: ChatGoogleGenerativeAI, tools: list):
    """Builds the decision_agent_node closure, bound to the given llm and MCP tools."""

    def decision_agent_node(state: AgentState) -> dict:
        system_prompt = _build_system_prompt(state)

        seed = [] if state.messages else [HumanMessage(content=state.raw_input)]
        conversation = [*state.messages, *seed]
        history = [SystemMessage(content=system_prompt), *conversation]
        just_executed_tools = bool(conversation) and isinstance(conversation[-1], ToolMessage)

        if not just_executed_tools:
            llm_with_tools = llm.bind_tools(tools)
            ai_message = llm_with_tools.invoke(history)

            if ai_message.tool_calls:
                return {
                    "messages": [*seed, ai_message],
                    "tools_triggered": [tc["name"] for tc in ai_message.tool_calls],
                    "thought_process": ai_message.content
                    or "Determined that an external tool call is required before finalizing the assessment.",
                }

        structured_llm = llm.with_structured_output(DecisionOutput)
        decision: DecisionOutput = structured_llm.invoke(history)

        return {
            "messages": seed,
            "thought_process": decision.thought_process,
            "tools_triggered": decision.tools_triggered,
            "final_urgency_assessment": decision.final_urgency_assessment,
            "recommended_action": decision.recommended_action,
            "summary_for_human_reviewer": decision.summary_for_human_reviewer,
        }

    return decision_agent_node


def should_continue(state: AgentState) -> str:
    if state.final_urgency_assessment:
        return "output_screening"
    if state.tools_triggered:
        return "execute_tools"
    return "output_screening"


# =============================================================================
# Output screening: HTTP call to services/output_screening (async, T3.micro)
# =============================================================================


def _context_severity_from_state(state: AgentState) -> str:
    """
    Maps this project's actual urgency vocabulary onto the screening
    service's context_severity input. final_urgency_assessment ("Low" /
    "Medium" / "High" / "Critical") is what's actually populated at this
    point in the graph -- state.risk_level is NOT: it's only set afterwards
    by routing_by_risk_level_node, which runs *after* output_screening, so
    it's always "" here and can't be used for this decision.
    """
    severity = state.final_urgency_assessment or state.distress_classification
    return severity.lower() if severity else "unknown"


async def _check_output_via_screening_service(incident_id: str, text: str, context_severity: str) -> dict:
    """
    Calls the standalone Output Screening microservice's POST /v1/screen
    (services/output_screening -- async, T3.micro-optimized: AWS Bedrock
    Guardrails + Claude Haiku), replacing the old guardrails_service
    /check/output call.

    This is a second, independent fail-safe layer on top of the one already
    inside the microservice itself (which has its own internal Bedrock
    timeout/retry): the microservice can still be unreachable at the network
    level -- a t3.micro can OOM/restart, or sit behind a corporate proxy
    (Zscaler) that blocks the connection outright, both observed on this
    exact project. A hard 3s client-side timeout means a dead microservice
    can never hang the graph; any failure here is treated as "not approved"
    so the crisis pipeline degrades to human review instead of either
    hanging or silently letting unchecked content through.
    """
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            response = await client.post(
                OUTPUT_SCREENING_URL,
                json={"event_id": incident_id, "raw_llm_output": text, "context_severity": context_severity},
            )
            response.raise_for_status()
            return response.json()
    except (httpx.TimeoutException, httpx.HTTPError, ValueError) as e:
        print(f"[EMERGENCY] Output Screening service unreachable: {e}")
        return {
            "event_id": incident_id,
            "is_safe": False,
            "action_approved": False,
            "screened_output": "",
            "guardrail_triggered": False,
            "failure_reason": "network_timeout" if isinstance(e, httpx.TimeoutException) else "proxy_error",
        }


async def output_screening_node(state: AgentState) -> dict:
    """
    Runs after decision_agent produces a final assessment (mirrors the n8n
    "Output Screening" node) and before "Routing by Risk Level" reads the
    graph's result. Delegates to the standalone Output Screening microservice
    over HTTP (see _check_output_via_screening_service). On failure, retries
    decision_agent up to MAX_OUTPUT_RETRIES times before giving up and
    routing straight to human_review.
    """
    text_to_screen = state.summary_for_human_reviewer or state.thought_process
    attempt_number = state.output_retry_count + 1
    context_severity = _context_severity_from_state(state)

    result = await _check_output_via_screening_service(state.incident_id, text_to_screen, context_severity)
    passed = bool(result.get("action_approved"))
    tags = ["guardrail_triggered"] if result.get("guardrail_triggered") else []

    screening_logs = [
        *state.screening_logs,
        {
            "attempt": attempt_number,
            "text_screened": text_to_screen,
            "context_severity": context_severity,
            "result": result,
            "passed": passed,
        },
    ]

    if passed:
        return {
            "screening_passed": True,
            "screening_tags": tags,
            "screening_reason": result.get("failure_reason", "none"),
            "screening_logs": screening_logs,
            "output_retry_count": 0,
            "output_screening_route": "routing_by_risk_level",
        }

    failure_reason = result.get("failure_reason", "proxy_error")

    if state.output_retry_count < MAX_OUTPUT_RETRIES:
        correction_message = HumanMessage(
            content=(
                "בדיקת הפלט (Output Screening) חסמה את התשובה הקודמת שלך ולא ניתן להעביר "
                f"אותה הלאה. סיבה: {failure_reason}. תקן/י את התשובה כך שלא תכלול תוכן חסום "
                "או רגיש ותתבסס אך ורק על ההקשר (RAG context) והעובדות שסופקו לך, ונסח/י "
                "מחדש הערכה סופית."
            )
        )
        return {
            "screening_passed": False,
            "screening_tags": tags,
            "screening_reason": failure_reason,
            "screening_logs": screening_logs,
            "output_retry_count": state.output_retry_count + 1,
            "output_screening_route": "decision_agent",
            "messages": [correction_message],
            "final_urgency_assessment": "",
        }

    return {
        "screening_passed": False,
        "screening_tags": tags,
        "screening_reason": failure_reason,
        "screening_logs": screening_logs,
        "risk_level": _fallback_risk_level(state),
        "output_screening_route": "human_review",
    }


def route_after_output_screening(state: AgentState) -> str:
    return state.output_screening_route or "routing_by_risk_level"


# =============================================================================
# Routing by risk level
# =============================================================================


def _fallback_risk_level(state: AgentState) -> str:
    urgency = (state.final_urgency_assessment or state.distress_classification or "").strip().lower()
    return URGENCY_TO_RISK_LEVEL.get(urgency, "medium")


def routing_by_risk_level_node(state: AgentState) -> dict:
    payload = {
        "incident_id": state.incident_id,
        "user_id": state.user_id,
        "distress_classification": state.distress_classification,
        "final_urgency_assessment": state.final_urgency_assessment,
        "summary_for_human_reviewer": state.summary_for_human_reviewer,
    }

    try:
        response = requests.post(ROUTING_WEBHOOK_URL, json=payload, timeout=10)
        response.raise_for_status()
        risk_level = (response.json().get("risk_level") or "").strip().lower()
        if risk_level not in ("high", "medium", "low"):
            raise ValueError(f"unexpected risk_level from routing workflow: {risk_level!r}")
    except (requests.exceptions.RequestException, ValueError) as e:
        print(f"[Routing Workflow Unreachable] {e}; falling back to local risk mapping.")
        risk_level = _fallback_risk_level(state)

    return {"risk_level": risk_level}


def route_by_risk_level(state: AgentState) -> str:
    return {"high": "immediate_alert", "medium": "human_review", "low": "log_and_close"}.get(
        state.risk_level, "human_review"
    )


# =============================================================================
# Terminal nodes
# =============================================================================


def immediate_alert_node(state: AgentState) -> dict:
    print(
        f"[EMERGENCY ALERT] incident={state.incident_id} user={state.user_id} "
        f"urgency={state.final_urgency_assessment} risk_level={state.risk_level} "
        f"summary={state.summary_for_human_reviewer}"
    )

    try:
        response = requests.post(
            ALERT_WEBHOOK_URL,
            json={
                "incident_id": state.incident_id,
                "user_id": state.user_id,
                "urgency_reason": state.summary_for_human_reviewer or state.thought_process,
                "risk_level": state.risk_level,
            },
            timeout=10,
        )
        response.raise_for_status()
        alert_status = "alert_sent"
    except requests.exceptions.RequestException as e:
        print(f"[Immediate Alert Error] could not reach alert webhook: {e}")
        alert_status = f"alert_failed: {e}"

    save_immediate_alert_record(
        incident_id=state.incident_id,
        user_id=state.user_id,
        risk_level=state.risk_level,
        alert_status=alert_status,
        urgency_reason=state.summary_for_human_reviewer or state.thought_process,
    )

    return {"recommended_action": alert_status}


def human_review_node(state: AgentState) -> dict:
    print(
        f"[HUMAN REVIEW] incident={state.incident_id} user={state.user_id} "
        f"risk_level={state.risk_level} summary={state.summary_for_human_reviewer} "
        f"screening_reason={state.screening_reason}"
    )

    try:
        response = requests.post(
            REVIEW_QUEUE_URL,
            json={
                "incident_id": state.incident_id,
                "user_id": state.user_id,
                "risk_level": state.risk_level,
                "summary_for_human_reviewer": state.summary_for_human_reviewer,
                "screening_logs": state.screening_logs,
            },
            timeout=10,
        )
        response.raise_for_status()
        review_status = "queued_for_human_review"
    except requests.exceptions.RequestException as e:
        print(f"[Human Review Error] could not reach review queue: {e}")
        review_status = f"review_queue_unreachable: {e}"

    return {"recommended_action": review_status}


INCIDENTS_TABLE_KEY = "incidents/incidents_table.xlsx"
INCIDENTS_TABLE_SHEET = "Incidents"
INCIDENTS_TABLE_COLUMNS = [
    "timestamp", "incident_id", "user_id", "risk_level", "distress_classification",
    "names", "addresses", "ages", "phone_numbers", "text_content",
    "summary_for_human_reviewer", "recommended_action", "final_urgency_assessment",
]


def _append_incident_to_s3_table(bucket: str, record: dict) -> None:
    s3 = boto3.client("s3", region_name=BEDROCK_AWS_REGION)

    try:
        download_buffer = io.BytesIO()
        s3.download_fileobj(bucket, INCIDENTS_TABLE_KEY, download_buffer)
        download_buffer.seek(0)
        wb = load_workbook(download_buffer)
        ws = wb[INCIDENTS_TABLE_SHEET] if INCIDENTS_TABLE_SHEET in wb.sheetnames else wb.active
    except (BotoCoreError, ClientError):
        wb = Workbook()
        ws = wb.active
        ws.title = INCIDENTS_TABLE_SHEET
        ws.append(INCIDENTS_TABLE_COLUMNS)

    ws.append([
        datetime.now().isoformat(timespec="seconds"),
        record.get("incident_id", ""),
        record.get("user_id", ""),
        record.get("risk_level", ""),
        record.get("distress_classification", ""),
        ", ".join(record.get("names") or []),
        ", ".join(record.get("addresses") or []),
        ", ".join(str(a) for a in (record.get("ages") or [])),
        ", ".join(record.get("phone_numbers") or []),
        record.get("raw_input", ""),
        record.get("summary_for_human_reviewer", ""),
        record.get("recommended_action", ""),
        record.get("final_urgency_assessment", ""),
    ])

    upload_buffer = io.BytesIO()
    wb.save(upload_buffer)
    upload_buffer.seek(0)
    s3.upload_fileobj(upload_buffer, bucket, INCIDENTS_TABLE_KEY)


def log_and_close_node(state: AgentState) -> dict:
    record = state.model_dump(exclude={"messages"})

    if S3_LOG_BUCKET:
        try:
            _append_incident_to_s3_table(S3_LOG_BUCKET, record)
            print(
                f"[Log & Close] incident={state.incident_id} appended to "
                f"s3://{S3_LOG_BUCKET}/{INCIDENTS_TABLE_KEY}"
            )
        except (BotoCoreError, ClientError) as e:
            print(f"[Log & Close] S3 archive failed for incident={state.incident_id}: {e}")
    else:
        print(
            f"[Log & Close] SAFESIGNAL_S3_BUCKET not configured -- skipping S3 archive "
            f"for incident={state.incident_id} (nothing else persists this incident)."
        )

    return {"recommended_action": "logged_and_closed"}


def build_graph(llm: ChatGoogleGenerativeAI, tools: list):
    graph = StateGraph(AgentState)
    graph.add_node("decision_agent", make_decision_agent_node(llm, tools))
    graph.add_node("execute_tools", ToolNode(tools))
    graph.add_node("output_screening", output_screening_node)
    graph.add_node("routing_by_risk_level", routing_by_risk_level_node)
    graph.add_node("immediate_alert", immediate_alert_node)
    graph.add_node("human_review", human_review_node)
    graph.add_node("log_and_close", log_and_close_node)

    graph.add_edge(START, "decision_agent")
    graph.add_conditional_edges(
        "decision_agent",
        should_continue,
        {"execute_tools": "execute_tools", "output_screening": "output_screening"},
    )
    graph.add_edge("execute_tools", "decision_agent")
    graph.add_conditional_edges(
        "output_screening",
        route_after_output_screening,
        {
            "decision_agent": "decision_agent",
            "human_review": "human_review",
            "routing_by_risk_level": "routing_by_risk_level",
        },
    )
    graph.add_conditional_edges(
        "routing_by_risk_level",
        route_by_risk_level,
        {
            "immediate_alert": "immediate_alert",
            "human_review": "human_review",
            "log_and_close": "log_and_close",
        },
    )
    graph.add_edge("immediate_alert", END)
    graph.add_edge("human_review", END)
    graph.add_edge("log_and_close", END)

    return graph.compile()


async def create_decision_agent(
    exit_stack: AsyncExitStack, model_name: str = MODEL_NAME, use_mock_tools: bool = False
):
    tools = build_mock_tools() if use_mock_tools else await load_all_mcp_tools(exit_stack)
    llm = ChatGoogleGenerativeAI(model=model_name, max_output_tokens=4096)
    return build_graph(llm, tools)
