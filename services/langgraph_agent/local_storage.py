import os
import threading
from datetime import datetime

from openpyxl import Workbook, load_workbook

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

ERROR_LOG_PATH = os.path.join(DATA_DIR, "error_log.xlsx")
ERROR_LOG_SHEET_NAME = "Errors"
ERROR_LOG_COLUMNS = [
    "timestamp", "source", "incident_id", "user_id", "error_message",
]

_lock = threading.Lock()


def _append_row(path: str, sheet_name: str, columns: list[str], row: list) -> None:
    """
    Shared append-with-init helper for all local Excel-backed tables in this
    module. Kept local-only (no S3/RDS) by design -- this data is sensitive
    enough that it stays on disk, not in the cloud, until a real encrypted
    store replaces it.
    """
    with _lock:
        if not os.path.exists(path):
            os.makedirs(os.path.dirname(path), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(columns)
            wb.save(path)

        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws.append(row)
        wb.save(path)


def save_error_record(
    source: str,
    error_message: str,
    incident_id: str = "",
    user_id: str = "",
    path: str = ERROR_LOG_PATH,
) -> None:
    """
    רשומת שגיאה מקומית (data/error_log.xlsx) - אותו דפוס בדיוק כמו
    review_queue/immediate_alert. יעד: מסך שגיאות עתידי ב-UI שיקרא מכאן
    דרך GET /api/v1/errors, כדי שכשל שרת (למשל n8n שלא הצליח להגיע ל-
    backend) לא ייבלע בשקט אלא יופיע למשתמש.
    """
    _append_row(path, ERROR_LOG_SHEET_NAME, ERROR_LOG_COLUMNS, [
        datetime.now().isoformat(timespec="seconds"),
        source,
        incident_id,
        user_id,
        error_message,
    ])


def get_error_records(path: str = ERROR_LOG_PATH) -> list[dict]:
    """קורא את data/error_log.xlsx בחזרה כרשימת dict-ים עבור מסך השגיאות ב-UI."""
    if not os.path.exists(path):
        return []

    with _lock:
        wb = load_workbook(path)
        ws = wb[ERROR_LOG_SHEET_NAME] if ERROR_LOG_SHEET_NAME in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

    return [dict(zip(ERROR_LOG_COLUMNS, row)) for row in rows]
