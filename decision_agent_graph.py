"""
LangGraph Decision Agent node + graph for the SafeSignal pipeline.

Loads tools from the two running FastMCP servers (alert_mcp.py on :8001,
rag_mcp.py on :8002), binds them to Gemini, and runs the full bottom half of
the pipeline diagram:

    decision_agent -> [tool call needed?] -> execute_tools -> decision_agent
    decision_agent -> output_screening
        -- blocked/hallucinated, retries left  --> back to decision_agent
        -- blocked/hallucinated, retries used up --> human_review
        -- passed                                --> routing_by_risk_level
    routing_by_risk_level
        -- risk_level == "high"   --> immediate_alert -> END
        -- risk_level == "medium" --> human_review -> END
        -- risk_level == "low"    --> log_and_close -> END

The first pass through decision_agent lets the model decide (via bound tools)
whether an external action is required. If not, it falls straight through to a
structured-output pass that produces the final DecisionOutput. If a tool call
was requested, ToolNode executes it and control loops back to decision_agent,
which then runs the structured-output pass over the tool results.

output_screening is a safety/quality gate on the agent's OWN generated text
(not a repeat of the earlier distress classification): AWS Bedrock Guardrails
via boto3 for banned words/profanity/PII (Hebrew + English), then Claude
Haiku on Bedrock to check the answer against the RAG context for
hallucinations. routing_by_risk_level then assigns the incident's risk_level
and fans out to the three terminal nodes.

Prerequisites: `alert_mcp.py` and `rag_mcp.py` must already be running (unless
using use_mock_tools=True), GOOGLE_API_KEY must be set, and AWS credentials
with Bedrock access must be configured (region via BEDROCK_AWS_REGION) for
output_screening/routing to do real checks instead of failing safe.
"""
import asyncio
import io
import json
import os
from contextlib import AsyncExitStack
from datetime import datetime

import boto3
import requests
from botocore.exceptions import BotoCoreError, ClientError
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()  # reads .env in the project root (git-ignored) into os.environ

from anthropic import AnthropicBedrock
from langchain_core.messages import HumanMessage, SystemMessage, ToolMessage
from langchain_core.tools import tool as lc_tool
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_mcp_adapters.tools import load_mcp_tools
from langgraph.graph import END, START, StateGraph
from langgraph.prebuilt import ToolNode
from mcp import ClientSession
from mcp.client.streamable_http import streamablehttp_client

from local_storage import save_immediate_alert_record
from schemas import AgentState, DecisionOutput

ALERT_MCP_URL = "http://localhost:8001/mcp"
RAG_MCP_URL = "http://localhost:8002/mcp"

MODEL_NAME = os.environ.get("DECISION_AGENT_MODEL", "gemini-2.5-flash")

# --- Output screening (Bedrock Guardrails + Claude Haiku hallucination check) ---
BEDROCK_AWS_REGION = os.environ.get("BEDROCK_AWS_REGION", "us-east-1")
# Provisioned separately in the AWS Bedrock console (word/PII/content filters);
# no ID configured means the guardrail step is skipped rather than failing.
BEDROCK_GUARDRAIL_ID = os.environ.get("BEDROCK_GUARDRAIL_ID", "")
BEDROCK_GUARDRAIL_VERSION = os.environ.get("BEDROCK_GUARDRAIL_VERSION", "DRAFT")
HALLUCINATION_MODEL_ID = os.environ.get(
    "BEDROCK_HALLUCINATION_MODEL_ID", "us.anthropic.claude-haiku-4-5-20251001-v1:0"
)
MAX_OUTPUT_RETRIES = 3

# --- Routing by risk level (n8n workflow, with a local fallback) ---
ROUTING_WEBHOOK_URL = os.environ.get("ROUTING_WEBHOOK_URL", "http://localhost:5678/webhook/risk-routing")
URGENCY_TO_RISK_LEVEL = {"critical": "high", "high": "high", "medium": "medium", "low": "low"}

# --- Terminal nodes ---
ALERT_WEBHOOK_URL = os.environ.get("ALERT_WEBHOOK_URL", "http://localhost:8000/api/v1/immediate-alert")
REVIEW_QUEUE_URL = os.environ.get("REVIEW_QUEUE_URL", "http://localhost:8000/api/v1/review-queue")

# Empty means "not provisioned yet" -- log_and_close_node then skips the S3 write
# instead of failing the whole (already-screened, low-risk) incident close-out.
S3_LOG_BUCKET = os.environ.get("SAFESIGNAL_S3_BUCKET", "")


async def load_all_mcp_tools(exit_stack: AsyncExitStack) -> list:
    """Connects to both FastMCP servers and returns their combined tool list."""
    tools = []
    for url in (ALERT_MCP_URL, RAG_MCP_URL):
        read, write, _ = await exit_stack.enter_async_context(streamablehttp_client(url))
        session = await exit_stack.enter_async_context(ClientSession(read, write))
        await session.initialize()
        tools.extend(await load_mcp_tools(session))
    return tools


def build_mock_tools() -> list:
    """
    Local stand-ins for the two MCP tools, with no network I/O -- same names and
    argument shapes as alert_mcp.py / rag_mcp.py, so the model's tool-calling decisions
    are unaffected. Used when the MCP servers aren't reachable (e.g. local testing
    without the two servers running, or a broken async MCP client environment).
    """

    @lc_tool
    def trigger_immediate_alert(incident_id: str, urgency_reason: str) -> str:
        """
        Trigger an immediate human/Amazon Polly voice alert for a high-risk incident.
        Call this only when the situation requires urgent human intervention (e.g.
        Critical or High urgency with an imminent safety risk).

        Args:
            incident_id: Unique identifier of the incident being escalated.
            urgency_reason: Short factual explanation of why immediate escalation is required.
        """
        return f"[MOCK] ALERT_SENT: simulated immediate alert for incident '{incident_id}' ({urgency_reason})."

    @lc_tool
    def query_rag_history(user_id: str, target_query: str) -> dict:
        """
        Fetch a specific user's historical distress patterns and past incident logs.
        Call this when the user's own history -- not just similar cases from the
        general knowledge base -- would materially change the urgency assessment.

        Args:
            user_id: Identifier of the user whose history should be searched.
            target_query: The specific question or topic to search the user's history for.
        """
        return {
            "mock": True,
            "user_id": user_id,
            "target_query": target_query,
            "history": ["[MOCK] No real history store connected -- this is simulated data."],
        }

    return [trigger_immediate_alert, query_rag_history]


def _build_system_prompt(state: AgentState) -> str:
    return (
        "You are the Decision Agent in SafeSignal, an emergency and distress triage "
        "system. Evaluate the incident context below and decide whether an external "
        "tool must be called before you can produce a final assessment.\n\n"
        f"Incident ID: {state.incident_id}\n"
        f"User ID: {state.user_id}\n"
        f"Raw input: {state.raw_input}\n"
        f"Automated distress classification: {state.distress_classification}\n"
        f"Passive RAG context already retrieved: {state.initial_rag_context}\n\n"
        "Available tools:\n"
        "- trigger_immediate_alert: use ONLY for Critical/High urgency situations that "
        "need immediate human/Amazon Polly voice intervention.\n"
        "- query_rag_history: use when this specific user's own historical incident "
        "pattern would materially change the urgency assessment and isn't already "
        "covered by the passive RAG context above.\n\n"
        "If no tool is needed, respond directly -- your final answer will be parsed "
        "into a structured assessment."
    )


def make_decision_agent_node(llm: ChatGoogleGenerativeAI, tools: list):
    """Builds the decision_agent_node closure, bound to the given llm and MCP tools."""

    def decision_agent_node(state: AgentState) -> dict:
        system_prompt = _build_system_prompt(state)

        # The user turn must live in state.messages (not just be passed transiently
        # to .invoke()), or on the loop-back pass the tool-call AIMessage ends up
        # with no preceding user turn in the persisted history -- which providers
        # with stricter function-calling validation (e.g. Gemini) reject outright:
        # "function call turn must come immediately after a user turn".
        seed = [] if state.messages else [HumanMessage(content=state.raw_input)]
        conversation = [*state.messages, *seed]
        history = [SystemMessage(content=system_prompt), *conversation]
        just_executed_tools = bool(conversation) and isinstance(conversation[-1], ToolMessage)

        if not just_executed_tools:
            llm_with_tools = llm.bind_tools(tools)
            ai_message = llm_with_tools.invoke(history)

            if ai_message.tool_calls:
                return {
                    "messages": [*seed, ai_message],
                    "tools_triggered": [tc["name"] for tc in ai_message.tool_calls],
                    "thought_process": ai_message.content
                    or "Determined that an external tool call is required before finalizing the assessment.",
                }

        structured_llm = llm.with_structured_output(DecisionOutput)
        decision: DecisionOutput = structured_llm.invoke(history)

        return {
            "messages": seed,
            "thought_process": decision.thought_process,
            "tools_triggered": decision.tools_triggered,
            "final_urgency_assessment": decision.final_urgency_assessment,
            "recommended_action": decision.recommended_action,
            "summary_for_human_reviewer": decision.summary_for_human_reviewer,
        }

    return decision_agent_node


def should_continue(state: AgentState) -> str:
    """Routes to execute_tools while a tool call is pending, otherwise proceeds to output screening."""
    if state.final_urgency_assessment:
        return "output_screening"
    if state.tools_triggered:
        return "execute_tools"
    return "output_screening"


# =============================================================================
# Output screening: AWS Bedrock Guardrails + Claude Haiku hallucination check
# =============================================================================

HALLUCINATION_SYSTEM_PROMPT = """את/ה בודק/ת בקרת-איכות על תשובה שנוצרה ע"י סוכן AI במערכת \
טריאז' מצוקה (SafeSignal). קיבלת כמה קטעי טקסט: "הודעת המשתמש המקורית" (raw input), \
"הסיווג האוטומטי" (distress classification) שנקבע לה קודם בפייפליין, ה"הקשר" (context) \
שאותר ממאגר הידע (RAG), "כלים שהופעלו בפועל" (tools actually executed) - רשימת שמות \
הכלים שהסוכן קרא להם בפועל במהלך הריצה, ו"פלט הסוכן" (agent output) - הערכת המצב הסופית \
שהסוכן ניסח.

תפקידך: לבדוק אם פלט הסוכן כולל טענות עובדתיות שאינן נתמכות ע"י אף אחד מהמקורות שסופקו \
(הודעת המשתמש המקורית, הסיווג האוטומטי, הקשר ה-RAG, או רשימת הכלים שהופעלו בפועל) - כלומר \
פרטים, המלצות פעולה קונקרטיות או קביעות עובדתיות שהסוכן "המציא" ואינן מבוססות על אף אחד \
מהם. ציטוט או תיאור של הודעת המשתמש המקורית, הפניה לסיווג האוטומטי שכבר נקבע, או טענה \
שפעולה/התראה מסוימת בוצעה/נשלחה - כאשר שם הכלי המתאים (למשל trigger_immediate_alert) \
מופיע ברשימת "כלים שהופעלו בפועל" - אינם הזיה, גם אם אינם מופיעים במאגר ה-RAG. טענה על \
פעולה שבוצעה כש*אין* כלי תואם ברשימה כן נחשבת הזיה.

אל תסמן כהזיה: ניסוח מחדש סביר, מסקנות לוגיות ישירות מההקשר, או שימוש בשיקול דעת מקצועי \
כללי (כמו "מומלץ ליצור קשר עם קו סיוע") שאינו סותר את ההקשר.

החזר/י JSON בלבד: hallucination_detected (bool) ו-reason (משפט קצר בעברית המסביר את ההחלטה).
"""

HALLUCINATION_SCHEMA = {
    "type": "object",
    "properties": {
        "hallucination_detected": {"type": "boolean"},
        "reason": {"type": "string"},
    },
    "required": ["hallucination_detected", "reason"],
    "additionalProperties": False,
}


def _run_bedrock_guardrails(text: str) -> dict:
    """
    Screens the Decision Agent's own generated text through an AWS Bedrock
    Guardrail (banned words, profanity, and sensitive PII such as credit-card
    numbers, in both Hebrew and English) via boto3's bedrock-runtime.apply_guardrail.

    Requires a Guardrail already provisioned in the AWS console, with its ID/
    version set via BEDROCK_GUARDRAIL_ID / BEDROCK_GUARDRAIL_VERSION. If unset
    (nothing provisioned yet) the check is explicitly skipped (not silently
    "passed"); if the AWS call itself fails, this fails *safe* by treating the
    text as blocked so it's retried/escalated rather than let through unchecked.
    """
    if not text or not text.strip():
        return {"blocked": False, "tags": [], "reason": "empty_output_nothing_to_screen"}

    if not BEDROCK_GUARDRAIL_ID:
        return {
            "blocked": False,
            "tags": [],
            "reason": "not_implemented: BEDROCK_GUARDRAIL_ID not configured -- guardrail check skipped",
        }

    try:
        client = boto3.client("bedrock-runtime", region_name=BEDROCK_AWS_REGION)
        response = client.apply_guardrail(
            guardrailIdentifier=BEDROCK_GUARDRAIL_ID,
            guardrailVersion=BEDROCK_GUARDRAIL_VERSION,
            source="OUTPUT",
            content=[{"text": {"text": text}}],
        )
    except (BotoCoreError, ClientError) as e:
        print(f"[Bedrock Guardrails Error] {e}")
        return {"blocked": True, "tags": ["guardrail_call_failed"], "reason": f"guardrail_call_failed: {e}"}

    action = response.get("action", "NONE")
    tags = sorted({key for assessment in response.get("assessments", []) for key in assessment.keys()})
    return {
        "blocked": action == "GUARDRAIL_INTERVENED",
        "tags": tags,
        "reason": f"bedrock_guardrail_action={action}",
    }


def _run_hallucination_check(
    agent_output: str,
    rag_context: str,
    raw_input: str = "",
    distress_classification: str = "",
    tools_triggered: list[str] | None = None,
) -> dict:
    """
    Uses Claude Haiku on Bedrock to compare the Decision Agent's final answer
    against everything the agent actually had grounds to work from -- the raw
    user message, the earlier automated distress classification, the RAG
    context gathered in the pipeline's Context Retrieval node, and the names
    of tools actually executed this run -- flagging claims the agent made
    that aren't grounded in any of those. Checking only against the RAG
    context (a handful of *similar* past cases, not the incident itself)
    previously caused false-positive hallucination flags any time the agent
    quoted the user's own message or cited the classification already
    established earlier in the run; omitting tools_triggered previously
    caused the same false-positive pattern for true statements like "an
    alert has been sent" whenever the matching tool (e.g.
    trigger_immediate_alert) really had been called -- the checker had no
    way to know that, since tool-execution results aren't part of the RAG/
    classification/raw-input sources it otherwise compares against. Fails
    *safe* on any error: treats the check as a hallucination so the graph
    retries/escalates instead of silently skipping the check.
    """
    if not agent_output or not agent_output.strip():
        return {"hallucination_detected": False, "reason": "empty_output_nothing_to_check"}

    try:
        client = AnthropicBedrock(aws_region=BEDROCK_AWS_REGION)
        user_content = (
            f"הודעת המשתמש המקורית:\n{raw_input or '(לא זמינה)'}\n\n"
            f"הסיווג האוטומטי שנקבע לפנייה:\n{distress_classification or '(לא זמין)'}\n\n"
            f"הקשר (RAG context) שאותר עבור הפנייה:\n{rag_context or '(אין הקשר זמין)'}\n\n"
            f"כלים שהופעלו בפועל:\n{tools_triggered or '(לא הופעל אף כלי)'}\n\n"
            f"פלט הסוכן לבדיקה:\n{agent_output}"
        )
        response = client.messages.create(
            model=HALLUCINATION_MODEL_ID,
            max_tokens=512,
            system=HALLUCINATION_SYSTEM_PROMPT,
            output_config={"format": {"type": "json_schema", "schema": HALLUCINATION_SCHEMA}},
            messages=[{"role": "user", "content": user_content}],
        )
        block = next(b for b in response.content if b.type == "text")
        parsed = json.loads(block.text)
        return {
            "hallucination_detected": bool(parsed["hallucination_detected"]),
            "reason": parsed["reason"],
        }
    except Exception as e:
        print(f"[Hallucination Check Error] {e}")
        return {"hallucination_detected": True, "reason": f"hallucination_check_failed: {e}"}


def output_screening_node(state: AgentState) -> dict:
    """
    Runs after decision_agent produces a final assessment (mirrors the n8n
    "Output Screening" node) and before "Routing by Risk Level" reads the
    graph's result. Two sequential checks on the agent's OWN generated text --
    not a repeat of the earlier distress classification:
      1. AWS Bedrock Guardrails -- banned words/profanity/PII (Hebrew + English).
      2. Claude Haiku on Bedrock -- hallucination check against the raw user
         input, the earlier distress classification, and the RAG context
         (skipped if step 1 already blocked the text).
    On failure, retries decision_agent up to MAX_OUTPUT_RETRIES times (with a
    correction message appended to state so the agent knows what to fix)
    before giving up and routing straight to human_review. Sets
    `output_screening_route` for this node's own conditional edge.
    """
    text_to_screen = state.summary_for_human_reviewer or state.thought_process
    attempt_number = state.output_retry_count + 1

    guardrail_result = _run_bedrock_guardrails(text_to_screen)
    if guardrail_result["blocked"]:
        hallucination_result = {"hallucination_detected": False, "reason": "skipped: blocked by guardrails"}
    else:
        hallucination_result = _run_hallucination_check(
            text_to_screen,
            state.initial_rag_context,
            state.raw_input,
            state.distress_classification,
            state.tools_triggered,
        )

    passed = not guardrail_result["blocked"] and not hallucination_result["hallucination_detected"]
    screening_logs = [
        *state.screening_logs,
        {
            "attempt": attempt_number,
            "text_screened": text_to_screen,
            "guardrail": guardrail_result,
            "hallucination_check": hallucination_result,
            "passed": passed,
        },
    ]

    if passed:
        return {
            "screening_passed": True,
            "screening_tags": guardrail_result["tags"],
            "screening_reason": hallucination_result["reason"],
            "screening_logs": screening_logs,
            "output_retry_count": 0,
            "output_screening_route": "routing_by_risk_level",
        }

    failure_reason = guardrail_result["reason"] if guardrail_result["blocked"] else hallucination_result["reason"]

    if state.output_retry_count < MAX_OUTPUT_RETRIES:
        correction_message = HumanMessage(
            content=(
                "בדיקת הפלט (Output Screening) חסמה את התשובה הקודמת שלך ולא ניתן להעביר "
                f"אותה הלאה. סיבה: {failure_reason}. תקן/י את התשובה כך שלא תכלול תוכן חסום "
                "או רגיש ותתבסס אך ורק על ההקשר (RAG context) והעובדות שסופקו לך, ונסח/י "
                "מחדש הערכה סופית."
            )
        )
        return {
            "screening_passed": False,
            "screening_tags": guardrail_result["tags"],
            "screening_reason": failure_reason,
            "screening_logs": screening_logs,
            "output_retry_count": state.output_retry_count + 1,
            "output_screening_route": "decision_agent",
            "messages": [correction_message],
            # Cleared so should_continue routes back through decision_agent's tool-
            # decision pass instead of re-screening the same still-blocked text.
            "final_urgency_assessment": "",
        }

    # Retries exhausted without a clean answer -- the agent's OUTPUT is still
    # unverified, but the urgency assessed earlier in the run is real and must
    # not be discarded. Fail safe: derive risk_level from that assessment (the
    # same local mapping used when the n8n routing workflow is unreachable)
    # instead of leaving it empty, so a critical case doesn't lose its priority
    # in the human-review queue just because the write-up failed screening.
    return {
        "screening_passed": False,
        "screening_tags": guardrail_result["tags"],
        "screening_reason": failure_reason,
        "screening_logs": screening_logs,
        "risk_level": _fallback_risk_level(state),
        "output_screening_route": "human_review",
    }


def route_after_output_screening(state: AgentState) -> str:
    return state.output_screening_route or "routing_by_risk_level"


# =============================================================================
# Routing by risk level
# =============================================================================


def _fallback_risk_level(state: AgentState) -> str:
    """Deterministic local mapping, used if the n8n routing workflow is unreachable."""
    urgency = (state.final_urgency_assessment or state.distress_classification or "").strip().lower()
    return URGENCY_TO_RISK_LEVEL.get(urgency, "medium")


def routing_by_risk_level_node(state: AgentState) -> dict:
    """
    Mirrors the n8n "Routing by Risk Level" node: decides the incident's final
    risk_level ("high" / "medium" / "low") from the conversation content and
    the earlier Distress Classification / Decision Agent urgency assessment.
    Prefers the n8n workflow (so ops can tune routing rules without a code
    deploy); falls back to a local deterministic mapping if n8n is unreachable
    or returns something unexpected, so a down workflow engine can't stall
    triage after the incident has already passed output screening.
    """
    payload = {
        "incident_id": state.incident_id,
        "user_id": state.user_id,
        "distress_classification": state.distress_classification,
        "final_urgency_assessment": state.final_urgency_assessment,
        "summary_for_human_reviewer": state.summary_for_human_reviewer,
    }

    try:
        response = requests.post(ROUTING_WEBHOOK_URL, json=payload, timeout=10)
        response.raise_for_status()
        risk_level = (response.json().get("risk_level") or "").strip().lower()
        if risk_level not in ("high", "medium", "low"):
            raise ValueError(f"unexpected risk_level from routing workflow: {risk_level!r}")
    except (requests.exceptions.RequestException, ValueError) as e:
        print(f"[Routing Workflow Unreachable] {e}; falling back to local risk mapping.")
        risk_level = _fallback_risk_level(state)

    return {"risk_level": risk_level}


def route_by_risk_level(state: AgentState) -> str:
    return {"high": "immediate_alert", "medium": "human_review", "low": "log_and_close"}.get(
        state.risk_level, "human_review"
    )


# =============================================================================
# Terminal nodes
# =============================================================================


def immediate_alert_node(state: AgentState) -> dict:
    """
    High-risk end node: pages a human responder immediately (n8n workflow ->
    Open WebUI dashboard + Amazon Polly voice call, per the pipeline diagram).
    A failed page must not raise and crash triage for a high-risk incident --
    it's logged loudly instead so the failure isn't silently lost.
    """
    print(
        f"[EMERGENCY ALERT] incident={state.incident_id} user={state.user_id} "
        f"urgency={state.final_urgency_assessment} risk_level={state.risk_level} "
        f"summary={state.summary_for_human_reviewer}"
    )

    try:
        response = requests.post(
            ALERT_WEBHOOK_URL,
            json={
                "incident_id": state.incident_id,
                "user_id": state.user_id,
                "urgency_reason": state.summary_for_human_reviewer or state.thought_process,
                "risk_level": state.risk_level,
            },
            timeout=10,
        )
        response.raise_for_status()
        alert_status = "alert_sent"
    except requests.exceptions.RequestException as e:
        print(f"[Immediate Alert Error] could not reach alert webhook: {e}")
        alert_status = f"alert_failed: {e}"

    # Until now this was the only terminal node with no persistent record --
    # review_queue/incident_log both write to data/*.xlsx, but the highest-
    # urgency cases only ever hit the console log above and vanished on
    # restart. Same local-Excel pattern as the other two terminal nodes.
    save_immediate_alert_record(
        incident_id=state.incident_id,
        user_id=state.user_id,
        risk_level=state.risk_level,
        alert_status=alert_status,
        urgency_reason=state.summary_for_human_reviewer or state.thought_process,
    )

    return {"recommended_action": alert_status}


def human_review_node(state: AgentState) -> dict:
    """
    Medium-risk (or screening-escalated) end node: queues the incident on the
    Open WebUI human-review dashboard instead of auto-resolving it.
    """
    print(
        f"[HUMAN REVIEW] incident={state.incident_id} user={state.user_id} "
        f"risk_level={state.risk_level} summary={state.summary_for_human_reviewer} "
        f"screening_reason={state.screening_reason}"
    )

    try:
        response = requests.post(
            REVIEW_QUEUE_URL,
            json={
                "incident_id": state.incident_id,
                "user_id": state.user_id,
                "risk_level": state.risk_level,
                "summary_for_human_reviewer": state.summary_for_human_reviewer,
                "screening_logs": state.screening_logs,
            },
            timeout=10,
        )
        response.raise_for_status()
        review_status = "queued_for_human_review"
    except requests.exceptions.RequestException as e:
        print(f"[Human Review Error] could not reach review queue: {e}")
        review_status = f"review_queue_unreachable: {e}"

    return {"recommended_action": review_status}


INCIDENTS_TABLE_KEY = "incidents/incidents_table.xlsx"
INCIDENTS_TABLE_SHEET = "Incidents"
INCIDENTS_TABLE_COLUMNS = [
    "timestamp", "incident_id", "user_id", "risk_level", "distress_classification",
    "names", "addresses", "ages", "phone_numbers", "text_content",
    "summary_for_human_reviewer", "recommended_action", "final_urgency_assessment",
]


def _append_incident_to_s3_table(bucket: str, record: dict) -> None:
    """
    Read-modify-write a single cumulative Excel workbook in S3 (one row per
    incident, one file total) instead of one JSON object per incident -- so
    the archive can be opened directly as a spreadsheet (2026-07-19 product
    decision) rather than requiring each incident's object to be opened one
    at a time. Downloads the existing workbook if present, appends a row,
    re-uploads. Not safe under concurrent writers (read-modify-write, no
    locking) -- same trade-off local_storage.py's _append_row already makes
    for the local Excel tables; acceptable here because this app has no
    concurrent traffic.
    """
    s3 = boto3.client("s3", region_name=BEDROCK_AWS_REGION)

    try:
        download_buffer = io.BytesIO()
        s3.download_fileobj(bucket, INCIDENTS_TABLE_KEY, download_buffer)
        download_buffer.seek(0)
        wb = load_workbook(download_buffer)
        ws = wb[INCIDENTS_TABLE_SHEET] if INCIDENTS_TABLE_SHEET in wb.sheetnames else wb.active
    except (BotoCoreError, ClientError):
        wb = Workbook()
        ws = wb.active
        ws.title = INCIDENTS_TABLE_SHEET
        ws.append(INCIDENTS_TABLE_COLUMNS)

    ws.append([
        datetime.now().isoformat(timespec="seconds"),
        record.get("incident_id", ""),
        record.get("user_id", ""),
        record.get("risk_level", ""),
        record.get("distress_classification", ""),
        ", ".join(record.get("names") or []),
        ", ".join(record.get("addresses") or []),
        ", ".join(str(a) for a in (record.get("ages") or [])),
        ", ".join(record.get("phone_numbers") or []),
        record.get("raw_input", ""),
        record.get("summary_for_human_reviewer", ""),
        record.get("recommended_action", ""),
        record.get("final_urgency_assessment", ""),
    ])

    upload_buffer = io.BytesIO()
    wb.save(upload_buffer)
    upload_buffer.seek(0)
    s3.upload_fileobj(upload_buffer, bucket, INCIDENTS_TABLE_KEY)


RAW_LOG_KEY = "raw/raw_messages_log.xlsx"
RAW_LOG_SHEET = "Raw Messages"
RAW_LOG_COLUMNS = ["timestamp", "incident_id", "user_id", "text_content"]


def _append_raw_message_to_s3_table(bucket: str, record: dict) -> bool:
    """
    Same read-modify-write pattern as _append_incident_to_s3_table, but for
    every ingested message before screening/classification -- one cumulative
    workbook instead of one raw/{user_id}/{incident_id}.json object per
    message (2026-07-20 product decision, mirrors the incidents table).

    Skips (no-op, returns False) if incident_id already has a row -- the
    Telegram poller re-fetches the same "last 20 messages" every scan with no
    upstream dedup, so without this check the same message gets re-appended
    on every cycle it's still within the last-20 window (2026-07-20 bug: 58
    of 80 rows were duplicates before this check existed).

    Uses the message's own send time (record["message_timestamp"], Telegram's
    unix-epoch msg.date) as the timestamp column when available, instead of
    "when this function happened to run" -- so the log reflects when messages
    were actually sent, not ingestion jitter. Falls back to now() for sources
    that don't carry a real send time (website/image/voice).

    Re-sorts the whole sheet by timestamp on every write instead of trusting
    append order. A single n8n scan posts ~20 of these concurrently and this
    function has no locking (same trade-off as _append_incident_to_s3_table),
    so whichever request's download/upload cycle happens to land last decides
    final row order -- sorting after every write keeps the file chronological
    regardless of arrival order (2026-07-20: requester wanted strict time
    ordering; per-item send order alone didn't guarantee it).
    """
    s3 = boto3.client("s3", region_name=BEDROCK_AWS_REGION)

    try:
        download_buffer = io.BytesIO()
        s3.download_fileobj(bucket, RAW_LOG_KEY, download_buffer)
        download_buffer.seek(0)
        wb = load_workbook(download_buffer)
        ws = wb[RAW_LOG_SHEET] if RAW_LOG_SHEET in wb.sheetnames else wb.active
    except (BotoCoreError, ClientError):
        wb = Workbook()
        ws = wb.active
        ws.title = RAW_LOG_SHEET
        ws.append(RAW_LOG_COLUMNS)

    existing_rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
    incident_id = record.get("incident_id", "")
    if incident_id and any(row[1] == incident_id for row in existing_rows):
        return False

    message_timestamp = record.get("message_timestamp")
    if message_timestamp:
        row_timestamp = datetime.fromtimestamp(message_timestamp).isoformat(timespec="seconds")
    else:
        row_timestamp = datetime.now().isoformat(timespec="seconds")

    existing_rows.append([
        row_timestamp,
        incident_id,
        record.get("user_id", ""),
        record.get("text_content", ""),
    ])
    existing_rows.sort(key=lambda row: row[0] or "")

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for row in existing_rows:
        ws.append(row)

    upload_buffer = io.BytesIO()
    wb.save(upload_buffer)
    upload_buffer.seek(0)
    s3.upload_fileobj(upload_buffer, bucket, RAW_LOG_KEY)
    return True


def log_raw_message(user_id: str, incident_id: str, text_content: str, message_timestamp=None) -> bool:
    """
    Best-effort append of one freshly-ingested message (any source: website/
    image/voice/Telegram) to the single cumulative raw-messages workbook in
    S3, before screening/classification. A storage hiccup here must not fail
    ingestion -- returns False instead of raising. Returns False without
    attempting a write if SAFESIGNAL_S3_BUCKET isn't configured, and False
    (not an error) if this incident_id was already logged in a prior scan.
    """
    if not S3_LOG_BUCKET:
        print("[Raw Log] SAFESIGNAL_S3_BUCKET not configured -- skipping raw S3 log.")
        return False

    try:
        appended = _append_raw_message_to_s3_table(
            S3_LOG_BUCKET,
            {
                "incident_id": incident_id,
                "user_id": user_id,
                "text_content": text_content,
                "message_timestamp": message_timestamp,
            },
        )
        if not appended:
            print(f"[Raw Log] incident={incident_id} already logged -- skipping duplicate.")
        return appended
    except (BotoCoreError, ClientError) as e:
        print(f"[Raw Log] S3 archive failed for incident={incident_id}: {e}")
        return False


def log_and_close_node(state: AgentState) -> dict:
    """
    Low-risk end node: appends the incident to a single cumulative S3 Excel
    table (the only place this incident gets persisted -- no RDS/Postgres
    instance is provisioned for this project, and the earlier local Excel
    fallback was deliberately removed 2026-07-19 in favor of one managed
    file), then closes out the graph run. The S3 write is best-effort -- a
    storage hiccup here must not turn an already-screened, genuinely
    low-risk incident into an unhandled exception.

    Deliberately includes identifying PII (names/addresses/ages/phone_numbers
    from Relevant Information Extraction, plus the raw message text) in the
    archived record -- an explicit product decision (2026-07-19) that every
    incident the system identifies as distress must be fully documented,
    overriding local_storage.py's separate, still-local-only PII policy for
    the earlier Screened Records table. Revisit if this bucket's access
    policy/encryption posture changes.
    """
    record = state.model_dump(exclude={"messages"})

    if S3_LOG_BUCKET:
        try:
            _append_incident_to_s3_table(S3_LOG_BUCKET, record)
            print(
                f"[Log & Close] incident={state.incident_id} appended to "
                f"s3://{S3_LOG_BUCKET}/{INCIDENTS_TABLE_KEY}"
            )
        except (BotoCoreError, ClientError) as e:
            print(f"[Log & Close] S3 archive failed for incident={state.incident_id}: {e}")
    else:
        print(
            f"[Log & Close] SAFESIGNAL_S3_BUCKET not configured -- skipping S3 archive "
            f"for incident={state.incident_id} (nothing else persists this incident)."
        )

    return {"recommended_action": "logged_and_closed"}


def build_graph(llm: ChatGoogleGenerativeAI, tools: list):
    graph = StateGraph(AgentState)
    graph.add_node("decision_agent", make_decision_agent_node(llm, tools))
    graph.add_node("execute_tools", ToolNode(tools))
    graph.add_node("output_screening", output_screening_node)
    graph.add_node("routing_by_risk_level", routing_by_risk_level_node)
    graph.add_node("immediate_alert", immediate_alert_node)
    graph.add_node("human_review", human_review_node)
    graph.add_node("log_and_close", log_and_close_node)

    graph.add_edge(START, "decision_agent")
    graph.add_conditional_edges(
        "decision_agent",
        should_continue,
        {"execute_tools": "execute_tools", "output_screening": "output_screening"},
    )
    graph.add_edge("execute_tools", "decision_agent")
    graph.add_conditional_edges(
        "output_screening",
        route_after_output_screening,
        {
            "decision_agent": "decision_agent",
            "human_review": "human_review",
            "routing_by_risk_level": "routing_by_risk_level",
        },
    )
    graph.add_conditional_edges(
        "routing_by_risk_level",
        route_by_risk_level,
        {
            "immediate_alert": "immediate_alert",
            "human_review": "human_review",
            "log_and_close": "log_and_close",
        },
    )
    graph.add_edge("immediate_alert", END)
    graph.add_edge("human_review", END)
    graph.add_edge("log_and_close", END)

    return graph.compile()


async def create_decision_agent(
    exit_stack: AsyncExitStack, model_name: str = MODEL_NAME, use_mock_tools: bool = False
):
    """
    Connects to both FastMCP servers (registering the sessions on `exit_stack` so
    they stay open for as long as the caller keeps the stack open) and returns the
    compiled decision-agent graph. Used both by main() below and by safesignal.py,
    which keeps `exit_stack` open for the lifetime of the FastAPI app.

    Pass `use_mock_tools=True` to skip the MCP servers entirely and run against the
    local stub tools from build_mock_tools() instead -- no network calls, no async
    MCP client session negotiation, `exit_stack` is left untouched.
    """
    tools = build_mock_tools() if use_mock_tools else await load_all_mcp_tools(exit_stack)
    llm = ChatGoogleGenerativeAI(model=model_name, max_output_tokens=4096)
    return build_graph(llm, tools)


async def main():
    async with AsyncExitStack() as exit_stack:
        graph = await create_decision_agent(exit_stack)

        initial_state = {
            "raw_input": "אני מרגיש מאוד לבד ומיואש בזמן האחרון, אין לי כוח להמשיך",
            "user_id": "user-1234",
            "incident_id": "incident-5678",
            "distress_classification": "high",
            "initial_rag_context": "מקרים דומים שאותרו במאגר הידע: תחושות בדידות ויאוש.",
        }

        final_state = await graph.ainvoke(initial_state)
        print(final_state)


if __name__ == "__main__":
    asyncio.run(main())
